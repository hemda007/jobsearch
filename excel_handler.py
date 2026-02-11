"""Reads/writes the Excel tracker file using openpyxl."""

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config

# Expected column headers (1-indexed)
EXPECTED_HEADERS = {
    1: "Job posting link",
    2: "JD Text",
    3: "Match %",
    4: "3 areas of improvement",
    5: "Referal profile 1",
    6: "Contextual intro",
    7: "Referal profile 2",
    8: "Contextual intro2",
    9: "Referal profile 3",
    10: "Contextual intro3",
}

COLUMN_WIDTHS = {
    1: 40,   # A - Job posting link
    2: 60,   # B - JD Text
    3: 10,   # C - Match %
    4: 50,   # D - 3 areas of improvement
    5: 40,   # E - Referal profile 1
    6: 50,   # F - Contextual intro
    7: 40,   # G - Referal profile 2
    8: 50,   # H - Contextual intro2
    9: 40,   # I - Referal profile 3
    10: 50,  # J - Contextual intro3
}

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _ensure_tracker_exists(tracker_path: str) -> None:
    """Create the tracker file with headers if it doesn't exist."""
    if os.path.exists(tracker_path):
        return

    os.makedirs(os.path.dirname(tracker_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    header_font = Font(bold=True)
    for col, header in EXPECTED_HEADERS.items():
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for col, width in COLUMN_WIDTHS.items():
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = width

    wb.save(tracker_path)
    print(f"Created new tracker file: {tracker_path}")


def _migrate_columns_if_needed(ws) -> bool:
    """
    Check if Column B is 'JD Text'. If not, insert it and shift data right.
    Returns True if migration was performed.
    """
    header_b = ws.cell(row=1, column=2).value
    if header_b and str(header_b).strip().lower() == "jd text":
        return False

    print("  Migrating Excel: inserting 'JD Text' column at B and shifting data right...")
    ws.insert_cols(2)
    ws.cell(row=1, column=2, value="JD Text")
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=2).alignment = Alignment(wrap_text=True)
    return True


def _apply_formatting(ws) -> None:
    """Apply column widths and text wrapping."""
    for col, width in COLUMN_WIDTHS.items():
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def read_unprocessed_rows(tracker_path: str = None) -> list[tuple[int, str, str]]:
    """
    Read the Excel tracker and find rows that need processing.

    Returns list of (row_number, job_link, jd_text) tuples.
    """
    if tracker_path is None:
        tracker_path = config.TRACKER_PATH

    _ensure_tracker_exists(tracker_path)

    # Early write-permission check — fail fast before expensive API calls
    try:
        with open(tracker_path, "a"):
            pass
    except PermissionError:
        raise PermissionError(
            f"Cannot write to {tracker_path} — the file is open in Excel or another application.\n"
            "Please close it and try again."
        )

    try:
        wb = load_workbook(tracker_path)
    except PermissionError:
        raise PermissionError(
            f"Cannot open {tracker_path} — the file may be open in another application.\n"
            "Please close it and try again."
        )

    ws = wb.active

    migrated = _migrate_columns_if_needed(ws)
    if migrated:
        _apply_formatting(ws)
        wb.save(tracker_path)
        print(f"  Column migration complete. Saved to {tracker_path}")
        print("  Please paste job description text into the new 'JD Text' column (B).")

    unprocessed = []
    for row in range(2, ws.max_row + 1):
        job_link = ws.cell(row=row, column=1).value
        jd_text = ws.cell(row=row, column=2).value
        match_pct = ws.cell(row=row, column=3).value

        if job_link and str(job_link).strip():
            if jd_text and str(jd_text).strip():
                if not match_pct or not str(match_pct).strip():
                    unprocessed.append((row, str(job_link).strip(), str(jd_text).strip()))
            else:
                print(f"  Row {row}: JD text is empty, skipping. Paste the job description in Column B.")

    wb.close()
    return unprocessed


class TrackerWriter:
    """Context manager that keeps the workbook open for batch writes."""

    def __init__(self, tracker_path: str = None):
        self.tracker_path = tracker_path or config.TRACKER_PATH
        self.wb = None
        self.ws = None

    def __enter__(self):
        try:
            self.wb = load_workbook(self.tracker_path)
        except PermissionError:
            raise PermissionError(
                f"Cannot open {self.tracker_path} — the file may be open in another application.\n"
                "Please close it and try again."
            )
        self.ws = self.wb.active
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.wb:
            self.wb.close()
        return False

    def write_results(self, row_number: int, results: dict) -> None:
        """Write processing results to a row and save immediately."""
        ws = self.ws

        # Column C: Match %
        match_pct = results.get("match_percentage", 0)
        cell_c = ws.cell(row=row_number, column=3, value=f"{match_pct}%")
        if match_pct >= 70:
            cell_c.fill = GREEN_FILL
        elif match_pct >= 50:
            cell_c.fill = YELLOW_FILL
        else:
            cell_c.fill = RED_FILL

        # Column D: 3 areas of improvement
        improvements = results.get("improvements", [])
        improvements_text = "\n".join(improvements) if isinstance(improvements, list) else str(improvements)
        ws.cell(row=row_number, column=4, value=improvements_text)

        # Columns E-J: Referrals and messages
        referrals = results.get("referrals", [])
        messages = results.get("messages", [])

        for i in range(3):
            ref_col = 5 + (i * 2)   # E=5, G=7, I=9
            msg_col = 6 + (i * 2)   # F=6, H=8, J=10

            if i < len(referrals):
                ref = referrals[i]
                ref_text = f"{ref['name']} | {ref['title']}\n{ref['url']}"
                ws.cell(row=row_number, column=ref_col, value=ref_text)
            else:
                ws.cell(row=row_number, column=ref_col, value="No profile found")

            if i < len(messages):
                ws.cell(row=row_number, column=msg_col, value=messages[i])
            else:
                ws.cell(row=row_number, column=msg_col, value="N/A")

        # Apply formatting to this row
        for col in range(1, 11):
            ws.cell(row=row_number, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        # Save after each row so progress isn't lost
        self.wb.save(self.tracker_path)
